import pyraw_nano
import openpyxl
import openpyxl.utils
import argparse
import numpy as np

def badMatchedPixels(D_true, D_estimated, error_thresh):

    mask = D_true > 0
    mask = mask > 0
    error = np.abs(np.round(D_estimated.astype(np.int32)) - D_true.astype(np.int32),)
    error[~mask] = 0

    perc = error > error_thresh
    BMP = np.sum(perc) / np.sum(mask)
    return BMP * 100.0


def main():
    ArgParser = argparse.ArgumentParser()
    ArgParser.add_argument('width', help='width')
    ArgParser.add_argument('height', help='height')
    ArgParser.add_argument('num_of_cams', help='num of cams')    
    ArgParser.add_argument('depthFromBlenderPath', help='depthFromBlenderPath')
    ArgParser.add_argument('depthFromGTParamsPath', help='depthFromGTParamsPath')
    ArgParser.add_argument('depthFromEstParamsPath', help='depthFromEstParamsPath')
    ArgParser.add_argument('resultsFilename', help='resultsFilename')
    Args = ArgParser.parse_args()

    W = int(Args.width)
    H = int(Args.height)
    pix_fmt = 'yuv420p16le'

    fnameX = 'v{view:01}_depth_'+str(W)+'x'+str(H)+'_'+str(pix_fmt)+'.yuv'
    fnameXX = 'v{view:02}_depth_'+str(W)+'x'+str(H)+'_'+str(pix_fmt)+'.yuv'
    

    # DM stands for DepthMap
    gtRenderDM    = Args.depthFromBlenderPath   + fnameXX
    calibParamsDM = Args.depthFromEstParamsPath + fnameX
    gtParamsDM    = Args.depthFromGTParamsPath  + fnameX

    ThresholdPerCent = [1, 2, 4]
    DepthMaxValue    = 2**16 - 1
    ThresholdValues  = [int(round(DepthMaxValue * (t/100))) for t in ThresholdPerCent]

    VIEWS = int(Args.num_of_cams)
    gtRenderDM_vs_calibParamsDM = []
    calibParamsDM_vs_gtParamsDM = []
    
    for view in range(VIEWS):
        gtRender    = pyraw_nano.imreadLuma(gtRenderDM   .format(view=view), W, H, 16, 420, 0)
        calibParams = pyraw_nano.imreadLuma(calibParamsDM.format(view=view), W, H, 16, 420, 0)
        gtParams    = pyraw_nano.imreadLuma(gtParamsDM   .format(view=view), W, H, 16, 420, 0)

        for threshold in ThresholdValues:
            gtRenderDM_vs_calibParamsDM.append([])
            gtRenderDM_vs_calibParamsDM[view].append(
                badMatchedPixels(gtRender, calibParams, threshold))
            
            calibParamsDM_vs_gtParamsDM.append([])
            calibParamsDM_vs_gtParamsDM[view].append(
                badMatchedPixels(calibParams, gtParams, threshold))


    wb = openpyxl.Workbook()
    ws = wb.create_sheet('gtRender_vs_calibParams')
    ws['A1'] = 'thresh'
    for i, t in enumerate(ThresholdPerCent):
            ws.cell(column=i+2, row=1, value=f"{t}%")
    
    for view in range(VIEWS):
        ws[f'A{view+2}'] = f'v{view}'
        for i, value in enumerate(gtRenderDM_vs_calibParamsDM[view]):
            ws.cell(column=i+2, row=view+2, value=value)

    ws.cell(column=1, row=VIEWS+2, value="AVERAGE")
    for i, value in enumerate(calibParamsDM_vs_gtParamsDM[view]):
         ColIdx = i+2
         ColLtr = openpyxl.utils.get_column_letter(ColIdx)
         ws.cell(column=ColIdx, row=VIEWS+2, value=f"=AVERAGE({ColLtr}2:{ColLtr}{VIEWS+1})")

    ws = wb.create_sheet('gtParams_vs_calibParams')
    ws['A1'] = 'thresh'
    for i, t in enumerate(ThresholdPerCent):
            ws.cell(column=i+2, row=1, value=f"{t}%")

    for view in range(VIEWS):
        ws[f'A{view+2}'] = f'v{view}'
        for i, value in enumerate(calibParamsDM_vs_gtParamsDM[view]):
            ws.cell(column=i+2, row=view+2, value=value)
    
    ws.cell(column=1, row=VIEWS+2, value="AVERAGE")
    for i, value in enumerate(calibParamsDM_vs_gtParamsDM[view]):
         ColIdx = i+2
         ColLtr = openpyxl.utils.get_column_letter(ColIdx)
         ws.cell(column=ColIdx, row=VIEWS+2, value=f"=AVERAGE({ColLtr}2:{ColLtr}{VIEWS+1})")

    del wb['Sheet']
    wb.save(Args.resultsFilename)


if __name__ == '__main__':
    main()